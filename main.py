# This is a sample Python script.
import os

import requests
from bs4 import BeautifulSoup
import json
import openpyxl
from decimal import Decimal, ROUND_HALF_UP


# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


def get_turret_info(tankInfo):
    if tankInfo['turrets'] is None:
        return None
    else:
        return tankInfo['turrets'][len(tankInfo['turrets']) - 1]


def gen_titles():
    ws.cell(1, 1).value = 'nation'
    ws.cell(1, 2).value = 'tier'
    ws.cell(1, 3).value = 'type'
    ws.cell(1, 4).value = 'name'
    ws.cell(1, 5).value = 'health'
    ws.cell(1, 6).value = 'weight'
    ws.cell(1, 7).value = 'ammo_rack_health'
    ws.cell(1, 8).value = 'ammo_rack_repair_health'
    ws.cell(1, 9).value = 'hull_armor_frot'
    ws.cell(1, 10).value = 'hull_armor_side'
    ws.cell(1, 11).value = 'hull_armor_rear'
    ws.cell(1, 12).value = 'fuel_tank_health'
    ws.cell(1, 13).value = 'fuel_tank_repair_health'
    ws.cell(1, 14).value = 'forward_speed'
    ws.cell(1, 15).value = 'reverse_speed'
    ws.cell(1, 16).value = 'xp_factor'
    ws.cell(1, 17).value = 'camo_price_factor'
    ws.cell(1, 18).value = 'camo_still'
    ws.cell(1, 19).value = 'camo_moving'
    ws.cell(1, 20).value = 'camo_fire_penalty'
    ws.cell(1, 21).value = 'camo_paint'
    ws.cell(1, 22).value = 'camo_net'
    ws.cell(1, 23).value = 'thrust-weight_ratio'
    ws.cell(1, 24).value = 'rotation_speed'
    ws.cell(1, 25).value = 'terrain_hard'
    ws.cell(1, 26).value = 'terrain_medium'
    ws.cell(1, 27).value = 'terrain_soft'
    ws.cell(1, 28).value = 'dispersion_movement'
    ws.cell(1, 29).value = 'dispersion_rotation'
    ws.cell(1, 30).value = 'chassis_health'
    ws.cell(1, 31).value = 'rotator_health'
    ws.cell(1, 32).value = 'view_range'
    ws.cell(1, 33).value = 'surveyor_health'
    ws.cell(1, 34).value = 'surveyor_repair_health'
    ws.cell(1, 35).value = 'turret_rotation_speed'
    ws.cell(1, 36).value = 'turret_armor_front'
    ws.cell(1, 37).value = 'turret_armor_side'
    ws.cell(1, 38).value = 'turret_armor_rear'
    ws.cell(1, 39).value = 'gun_health'
    ws.cell(1, 40).value = 'dpm'
    ws.cell(1, 41).value = 'gun_elevation'
    ws.cell(1, 42).value = 'gun_depression'
    ws.cell(1, 43).value = 'gun_rotation_speed'
    ws.cell(1, 44).value = 'gun_max_ammo'
    ws.cell(1, 45).value = 'gun_reload_time'
    ws.cell(1, 46).value = 'gun_aim_time'
    ws.cell(1, 47).value = 'gun_max_ammo'
    ws.cell(1, 48).value = 'gun_dispersion'
    ws.cell(1, 49).value = 'gun_dispersion_rotation'
    ws.cell(1, 50).value = 'gun_dispersion_firing'
    ws.cell(1, 51).value = 'gun_dispersion_damaged'
    ws.cell(1, 52).value = 'gun_clip_size'
    ws.cell(1, 53).value = 'gun_clip_reload'
    ws.cell(1, 54).value = 'gun_burst_size'
    ws.cell(1, 55).value = 'gun_burst_reload'
    ws.cell(1, 56).value = 'autoreload_time'
    ws.cell(1, 57).value = 'autoreload_fraction'
    ws.cell(1, 58).value = 'dual_reload_time'
    ws.cell(1, 59).value = 'dual_rate_time'
    ws.cell(1, 60).value = 'dual_charge_time'
    ws.cell(1, 61).value = 'dual_charge_threshold'
    ws.cell(1, 62).value = 'dual_charge_cancel_time'
    ws.cell(1, 63).value = 'dual_pre_charge_indication'
    ws.cell(1, 64).value = 'dual_reload_lock_time'
    ws.cell(1, 65).value = 'dual_after_shot_delay'
    ws.cell(1, 66).value = 'camo_fire_penalty'
    ws.cell(1, 67).value = 'caliber'
    ws.cell(1, 68).value = 'shell1_penetration'
    ws.cell(1, 69).value = 'shell2_penetration'
    ws.cell(1, 70).value = 'shell3_penetration'
    ws.cell(1, 71).value = 'shell1_damage'
    ws.cell(1, 72).value = 'shell2_damage'
    ws.cell(1, 73).value = 'shell3_damage'
    ws.cell(1, 74).value = 'shell1_speed'
    ws.cell(1, 75).value = 'shell2_speed'
    ws.cell(1, 76).value = 'shell3_speed'
    ws.cell(1, 77).value = 'shell1_module_damage'
    ws.cell(1, 78).value = 'shell2_module_damage'
    ws.cell(1, 79).value = 'shell3_module_damage'
    ws.cell(1, 80).value = 'shell1_explosion_radius'
    ws.cell(1, 81).value = 'shell2_explosion_radius'
    ws.cell(1, 82).value = 'shell3_explosion_radius'
    ws.cell(1, 83).value = 'engine_power'
    ws.cell(1, 84).value = 'engine_fire_chance'
    ws.cell(1, 85).value = 'engine_health'
    ws.cell(1, 86).value = 'engine_repair_health'
    ws.cell(1, 87).value = 'radio_health'
    ws.cell(1, 88).value = 'radio_repair_health'
    ws.cell(1, 89).value = 'radio_range'


def get_ratio_info():
    return tankInfo['radios'][len(tankInfo['radios']) - 1]


def get_chassis_info():
    if tankInfo['chassis'] is None:
        return None
    else:
        return tankInfo['chassis'][len(tankInfo['chassis']) - 1]


def get_gun_info():
    return tankInfo['guns'][len(tankInfo['guns']) - 1]


def get_engine_info():
    return tankInfo['engines'][len(tankInfo['engines']) - 1]


def get_wheel_info():
    if tankInfo['wheels'] is None:
        return None
    else:
        return tankInfo['wheels'][len(tankInfo['wheels']) - 1]


def get_weight():
    return tankInfo['weight'] + tankInfo['fuel_tank_weight'] + turret['weight'] + radio['weight'] + gun['weight'] + \
           engine[
               'weight'] + (chassis['weight'] if not isWheel else wheel['weight'])


def round_half_up(float_value):
    return Decimal(str(float_value)).quantize(Decimal("0.00"), rounding=ROUND_HALF_UP)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    url = 'https://tanks.gg/api/list'
    urlTankPrefix = 'https://tanks.gg/api/v11210/tank/'
    res = requests.get(url)
    tanks = json.loads(res.text)['tanks']
    tank103 = json.loads(requests.get('https://tanks.gg/api/v11210/tank/m103').text)
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('v11210', 0)
    gen_titles()

    tankInfo = tank103['tank']
    turret = get_turret_info(tankInfo)
    radio = get_ratio_info()
    chassis = get_chassis_info()
    wheel = get_wheel_info()
    isWheel = wheel is not None
    gun = get_gun_info()
    engine = get_engine_info()
    gunShell = []
    shells = tankInfo['shells']
    crewSkillFactor = 0.95877277
    for shell in shells:
        if shell['gun_id'] == gun['id']:
            gunShell.append(shell)
    ws.cell(2, 1).value = tankInfo['nation']
    ws.cell(2, 2).value = tankInfo['tier']
    ws.cell(2, 3).value = tankInfo['type']
    ws.cell(2, 4).value = tankInfo['name']
    ws.cell(2, 5).value = tankInfo['health'] + turret['health']
    weight = get_weight()
    ws.cell(2, 6).value = weight
    ws.cell(2, 7).value = tankInfo['ammo_rack_health']
    ws.cell(2, 8).value = tankInfo['ammo_rack_repair_price']
    ws.cell(2, 9).value = tankInfo['armor_front']
    ws.cell(2, 10).value = tankInfo['armor_side']
    ws.cell(2, 11).value = tankInfo['armor_rear']
    ws.cell(2, 12).value = tankInfo['fuel_tank_health']
    ws.cell(2, 13).value = tankInfo['fuel_tank_repair_health']
    ws.cell(2, 14).value = tankInfo['forward_speed']
    ws.cell(2, 15).value = tankInfo['reverse_speed']
    ws.cell(2, 16).value = tankInfo['xp_factor']
    ws.cell(2, 17).value = tankInfo['camo_price_factor']
    ws.cell(2, 18).value = tankInfo['camo_still']
    ws.cell(2, 19).value = tankInfo['camo_moving']
    ws.cell(2, 20).value = tankInfo['camo_fire_penalty']
    ws.cell(2, 21).value = tankInfo['camo_paint']
    ws.cell(2, 22).value = tankInfo['camo_net']
    ws.cell(2, 23).value = round_half_up(engine['power'] / weight)
    if isWheel:
        ws.cell(2, 24).value = 19
    else:
        ws.cell(2, 24).value = chassis['rotation_speed']
        ws.cell(2, 25).value = chassis['terrain_hard']
        ws.cell(2, 26).value = chassis['terrain_medium']
        ws.cell(2, 27).value = chassis['terrain_soft']
        ws.cell(2, 28).value = chassis['dispersion_movement']
        ws.cell(2, 29).value = chassis['dispersion_rotation']
        ws.cell(2, 30).value = chassis['health']
    if turret is not None:
        ws.cell(2, 31).value = turret['rotator_health']
        ws.cell(2, 32).value = turret['view_range']
        ws.cell(2, 33).value = turret['surveyor_health']
        ws.cell(2, 34).value = turret['surveyor_repair_health']
        ws.cell(2, 35).value = turret['rotation_speed']
        ws.cell(2, 36).value = turret['armor_front']
        ws.cell(2, 37).value = turret['armor_side']
        ws.cell(2, 38).value = turret['armor_rear']
    ws.cell(2, 39).value = gun['health']
    ws.cell(2, 40).value = round_half_up(60 / gun['reload_time'] / crewSkillFactor * gunShell[0]['damage'])
    ws.cell(2, 41).value = gun['elevation']
    ws.cell(2, 42).value = gun['depression']
    ws.cell(2, 43).value = gun['rotation_speed']
    ws.cell(2, 44).value = gun['max_ammo']
    ws.cell(2, 45).value = round_half_up(gun['reload_time'] * crewSkillFactor)
    ws.cell(2, 46).value = round_half_up(gun['aim_time'] * crewSkillFactor)
    ws.cell(2, 47).value = gun['max_ammo']
    ws.cell(2, 48).value = round_half_up(gun['dispersion'] * crewSkillFactor)
    ws.cell(2, 49).value = gun['dispersion_rotation']
    ws.cell(2, 50).value = gun['dispersion_firing']
    ws.cell(2, 51).value = gun['dispersion_damaged']
    ws.cell(2, 52).value = gun['clip_size']
    ws.cell(2, 53).value = gun['clip_reload']
    ws.cell(2, 54).value = gun['burst_size']
    ws.cell(2, 55).value = gun['burst_reload']
    ws.cell(2, 56).value = gun['autoreload_time']
    ws.cell(2, 57).value = gun['autoreload_fraction']
    ws.cell(2, 58).value = gun['dual_reload_time']
    ws.cell(2, 59).value = gun['dual_rate_time']
    ws.cell(2, 60).value = gun['dual_charge_time']
    ws.cell(2, 61).value = gun['dual_charge_threshold']
    ws.cell(2, 62).value = gun['dual_charge_cancel_time']
    ws.cell(2, 63).value = gun['dual_pre_charge_indication']
    ws.cell(2, 64).value = gun['dual_reload_lock_time']
    ws.cell(2, 65).value = gun['dual_after_shot_delay']
    ws.cell(2, 66).value = gun['camo_fire_penalty']
    ws.cell(2, 67).value = gunShell[0]['caliber']
    shellIndex = 0
    for shell in gunShell:
        ws.cell(2, 68 + shellIndex).value = gunShell[shellIndex]['penetration']
        ws.cell(2, 71 + shellIndex).value = gunShell[shellIndex]['damage']
        ws.cell(2, 74 + shellIndex).value = gunShell[shellIndex]['speed']
        ws.cell(2, 77 + shellIndex).value = gunShell[shellIndex]['module_damage']
        ws.cell(2, 80 + shellIndex).value = gunShell[shellIndex]['explosion_radius']
        shellIndex = shellIndex + 1
    ws.cell(2, 83).value = engine['power']
    ws.cell(2, 84).value = engine['fire_chance']
    ws.cell(2, 85).value = engine['health']
    ws.cell(2, 86).value = engine['repair_health']
    ws.cell(2, 87).value = radio['health']
    ws.cell(2, 88).value = radio['repair_health']
    ws.cell(2, 89).value = radio['range']
    wb.save('m103.xlsx')

# for tank in tanks:
#     urlTank = urlTankPrefix + tank['slug']
#     tankInfo = requests.get(urlTank)
#     wb = openpyxl.Workbook()
#     ws = wb.create_sheet('v11210', 0)
#
#     print(urlTank)
# print_hi('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
